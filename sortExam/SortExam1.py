# coding=utf-8
import sys, os
# 解析 xls
import xlrd
# 解析 xlsx
import openpyxl
import math
from datetime import datetime, timedelta

# ///////////////////////////////////////////////////
# 排布规则:
# 能排一天的不跨天
# 一周能考完的不跨周
# openpyxl: pip install openpyxl
# 使用 python3 运行脚本
# ///////////////////////////////////////////////////

# 每日场次 python 2.7 整数相除为整数，没有小数>>>使用 python3
s_dayNum = 6
# 每场人数
s_sessionNum = 100
# 第几天有间隔，间隔几天
s_spaceIndex = 6
s_spaceNum = 3

# 时间排布
s_begin_year = 2024
s_begin_month = 6
s_begin_day = 3
s_times = ["08:30", "10:30", "12:30", "14:30", "16:30", "18:30"]

# 表结构
# 考点编号*,考点名称*,时间单元编号*,考试开始时间*,考场编号*,考场名称*,座位号*,考生学号*,考生姓名*,课程编号*,课程名称*,试卷号*
s_keyIndex = 2  # 索引字段-学号
s_typeIndex = 4  # 类型索引-课程编号
#先排索引小的
s_sortIndex = 7 # 排序索引

# 开始时间(周几)
s_beginWeekDay = 0
# 开始周考试天数
s_beginStageSessionNum = 0
# 表头
s_head = ""


# 待处理文件
s_originalFile = None

# 考生
s_personKeys = []


# 获取指定年月日相差天数的，年月日
def calculate_new_date(year, month, day, days_difference):    
    # 构造日期对象
    current_date = datetime(year, month, day)
    # 计算相差指定天数后的日期
    new_date = current_date + timedelta(days=days_difference)
    return new_date.year, new_date.month, new_date.day

# 人
def getTime(session):
    global s_begin_year
    global s_begin_month
    global s_begin_day
    global s_dayNum
    global s_times
    global s_spaceIndex
    global s_spaceNum
    global s_beginStageSessionNum

    # bug-间隔天数没跨月，先这样处理了
    _beginDay = s_begin_day
    if session > s_beginStageSessionNum:
        _beginDay = _beginDay + s_spaceNum

    year, month, day = calculate_new_date(s_begin_year, s_begin_month, _beginDay, (session - 1) / s_dayNum)
    
    index = session % s_dayNum
    if index == 0:
        index = len(s_times)
    index = index - 1

    if month < 10:
        month = str(month)
    if day < 10:
        day = str(day)
    return str(year) + "/" + str(month) + "/" + str(day) + " " + s_times[int(index)], str(session)

class Person:
    def __init__(self, _key, _type):
        self.m_key = _key
        self.m_type = _type
        self.sortIndex = None
        self.rows = []

    def getKey(self):
        return self.m_key

    def addRow(self, row, _type):
        self.rows.append(row)
        if self.sortIndex is None:
            self.sortIndex = getSortIndex(row)
        elif self.sortIndex != getSortIndex(row):
            print("错误：一个人有多个 index", self.m_key)
            exit()

    def getRowNum(self):
        return len(self.rows)

    def getRows(self):
        return self.rows

    def setSession(self, session):
        for row in self.rows:
            time, tag = getTime(session)
            row.addValues(session, time, tag)
            session += 1
    
    def getSortIndex(self):
        return self.sortIndex


# 寻找已有的 person
def getPerson(_key, personArr):
    for person in personArr:
        if person.getKey() == _key:
            return person
    return None


# 获取行的 key, type
def getKeyAndType(row):
    global s_keyIndex
    global s_typeIndex

    return row.values[s_keyIndex], row.values[s_typeIndex]

# 获取排序 index
def getSortIndex(row):
    global s_sortIndex

    return row.values[s_sortIndex]

# ///////// 读 excel 内容 ///////////////
# excel 行数据
class ExcelRow:
    # row xlsx 读得的值
    def __init__(self, row = None):
        self.values = []
        if row:
            for cell in row:
                self.values.append(cell.value)
        # print(row[0].value)

    # xls 读取的值
    def addValue(self, value):
        # 数字单元格的结果后面多了一个.0，这是因为xlrd默认将数字单元格的数据类型解析为浮点数（float），即使实际上是整数
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        self.values.append(value)

    # 加：场次，时间，时间编号
    def addValues(self, session, time, tag):
        self.values.append(session)
        self.values.append(time)
        self.timeIndex = len(self.values) - 1
        self.values.append(tag)

    def getRowContent(self):
        content = ""
        for value in self.values:
            content += str(value) + ","
        return content.replace(",", "")

    def getTimeIndex(self):
        return self.timeIndex

# 读原文件 xls
def readXls():
    global s_originalFile

    headRow = None
    personArr = []
    _workbook = xlrd.open_workbook(s_originalFile)
    _sheets = _workbook.sheets()
    if len(_sheets) == 0:
        print(u"excel 文件内容为空！！")
        exit()
    if len(_sheets) > 1:
        print(u"警告！！存在多个sheet，只处理第一个sheet，如后面的 sheet 要处理，都拷贝到第一个 sheet!!!")
        print(u"警告！！存在多个sheet，只处理第一个sheet，如后面的 sheet 要处理，都拷贝到第一个 sheet!!!")
        print(u"警告！！存在多个sheet，只处理第一个sheet，如后面的 sheet 要处理，都拷贝到第一个 sheet!!!")
    # sheet = _workbook.sheet_by_index(0)
    sheet = _sheets[0]
    row_index = 0
    for row in range(sheet.nrows):
        if row_index == 0:
            headRow = ExcelRow()
            for col in range(sheet.ncols):
                value = sheet.cell_value(row, col)
                headRow.addValue(value)
            headRow.addValues("session", "time", "time_tag")
        else:
            excel_row = ExcelRow()
            for col in range(sheet.ncols):
                value = sheet.cell_value(row, col)
                excel_row.addValue(value)
            _key, _type = getKeyAndType(excel_row)
            person = getPerson(_key, personArr)
            if person is None:
                s_personKeys.append(_key)
                person = Person(_key, _type)
                personArr.append(person)
            person.addRow(excel_row, _type)
        row_index += 1

    print(u"场数：{}".format(row_index - 1))
    print(u"人数：{}".format(len(personArr)))
    return headRow, personArr


# 读原文件 xlsx
def readXlsx():
    global s_originalFile

    headRow = None
    personArr = []
    workbook = openpyxl.load_workbook(s_originalFile)
    if len(workbook.sheetnames) == 0:
        print(u"excel 文件内容为空！！")
        exit()
    elif len(workbook.sheetnames) > 1:
        print(u"警告！！存在多个sheet，只处理第一个sheet，如后面的 sheet 要处理，都拷贝到第一个 sheet!!!")
        print(u"警告！！存在多个sheet，只处理第一个sheet，如后面的 sheet 要处理，都拷贝到第一个 sheet!!!")
        print(u"警告！！存在多个sheet，只处理第一个sheet，如后面的 sheet 要处理，都拷贝到第一个 sheet!!!")
    sheet = workbook.worksheets[0]
    row_index = 0
    for _row in sheet.iter_rows():
        if row_index == 0:
            headRow = ExcelRow(_row)
            headRow.addValues("session", "time", "time_tag")
        else:
            row = ExcelRow(_row)
            _key, _type = getKeyAndType(row)
            person = getPerson(_key, personArr)
            if person is None:
                s_personKeys.append(_key)
                person = Person(_key, _type)
                personArr.append(person)
            person.addRow(row, _type)
        row_index += 1

    print(u"场数：{}".format(row_index - 1))
    print(u"人数：{}".format(len(personArr)))
    return headRow, personArr


# 场次多的人排前面
def sortPersons(personArr):
    print(u"index小的优先排前面，其次场次多的排前面")
    length = len(personArr)
    for i in range(0, length):
        for j in range(i + 1, length):
            if personArr[j].getSortIndex() < personArr[i].getSortIndex():
                temp = personArr[i]
                personArr[i] = personArr[j]
                personArr[j] = temp
            elif personArr[j].getSortIndex() == personArr[i].getSortIndex() and personArr[j].getRowNum() > personArr[i].getRowNum():
                temp = personArr[i]
                personArr[i] = personArr[j]
                personArr[j] = temp
    print(u"排序完毕")


# 第一阶段能进行的场次
def setFirstStageExamNum():
    global s_dayNum
    global s_beginWeekDay
    global s_spaceIndex
    global s_beginStageSessionNum

    # 第一阶段场次
    s_beginStageSessionNum = (s_spaceIndex - 1) * s_dayNum


# 安排场次是否超天或超周
def isSessionOverDayOrStage(session, person):
    global s_dayNum
    global s_beginStageSessionNum
    
    _personLineNum = person.getRowNum()
    ### 是否超天
    # 需要用到考试天数
    _needDay = math.ceil(_personLineNum / s_dayNum)
    # 耗费天数
    _firstDayLeft = 0  # 第一天剩余场次
    if session < s_dayNum:
        _firstDayLeft = s_dayNum - session + 1
    else:
        if session % s_dayNum == 0:
            _firstDayLeft = 1
        else:
            _firstDayLeft = s_dayNum - (session % s_dayNum) + 1

    if _personLineNum <= _firstDayLeft:
        return False
    _costDay = math.ceil((_personLineNum - _firstDayLeft) / s_dayNum) + 1

    if _costDay > _needDay:
        return True
    ### 不在第一阶段内，不超阶段
    if session > s_beginStageSessionNum:
        return False
    # 第一阶段够用
    if s_beginStageSessionNum - session + 1 > _personLineNum:
        return False
    return True


# 获取当前场次
def getNowSession(sessionInfo, person):
    global s_dayNum
    global s_sessionNum

    _session = 0
    while True:
        _session += 1
        if str(_session) in sessionInfo and sessionInfo[str(_session)] >= s_sessionNum:
            # 场次已满
            continue
        if isSessionOverDayOrStage(_session, person):
            # 超天或超第一阶段
            continue
        if not isNextSessionEnough(_session, sessionInfo, person):
            # 后续 session 排满了
            continue
        return _session


# 后续场次是否足够
def isNextSessionEnough(session, sessionInfo, person):
    global s_sessionNum

    for i in range(session, session + person.getRowNum()):
        _sKey = str(i)
        if _sKey in sessionInfo and sessionInfo[_sKey] >= s_sessionNum:
            return False
    return True


# 记录 session 数量
def addSessionInfo(sessionInfo, session, person):
    for i in range(session, session + person.getRowNum()):
        _sKey = str(i)
        if _sKey not in sessionInfo:
            sessionInfo[_sKey] = 1
        else:
            sessionInfo[_sKey] += 1


# 开始排序，获得结果
def generateResult(headRow, personArr, resultFile):
    global s_head

    print(u"获取结果内容")
    # 场次信息：{"1":100} 场次对应次数
    _sessionInfo = {}
    # 已排的人
    _sortKeys = []

    setFirstStageExamNum()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    print(u"写入头.....")
    for index, value in enumerate(headRow.values):
        # row, column 从 1 开始
        cell = sheet.cell(row=1, column=index + 1)
        cell.value = value

    print(u"写入详细.....")
    # 上面加了头，从第二行开始写
    _rowIndex = 2
    personIndex = 0
    for person in personArr:
        _personKey = person.getKey()
        # 获取从哪场开始排
        session = getNowSession(_sessionInfo, person)
        addSessionInfo(_sessionInfo, session, person)
        _sortKeys.append(_personKey)
        person.setSession(session)
        personIndex += 1
        for index, row in enumerate(person.getRows()):
            for _index, value in enumerate(row.values):
                # row, column 从 1 开始
                cell = sheet.cell(row=_rowIndex, column=_index + 1)
                if _index == row.getTimeIndex():
                    # 设置单元格的时间格式
                    cell.number_format = 'yyyy/m/d hh:mm'
                else:
                    # 科学计数法问题
                    # 设置单元格格式
                    cell.number_format = '0'  # 或者使用 '0.00' 等形式，确保数字以常规格式显示，而非科学计数法
                cell.value = value
            _rowIndex += 1

    workbook.save(resultFile)
    workbook.close()


if __name__ == '__main__':
    print("__main__>>>>>>>>>>>>>>>>>>>>>>>>")
    if sys.version_info.major < 3:
        print(u"使用 python3 执行脚本！！！！！")
        exit()
    
    if len(sys.argv) < 2:
        print(u"请输入要排序的文件名!")
        exit()
    s_originalFile = sys.argv[1]
    if not os.path.isfile(s_originalFile):
        print(u"输入的文件不存在！")
        exit()
        
    headRow = None
    personArr = None
    _, file_ext = os.path.splitext(s_originalFile)
    isXlsx = False
    if file_ext == ".xlsx":
        headRow, personArr = readXlsx()
    elif file_ext == ".xls":
        headRow, personArr = readXls()
    else:
        print(u"只支持排序 .xlsx 和 .xls 文件")
        exit()

    # 开始时间(周几)
    dateObject = datetime(s_begin_year, s_begin_month, s_begin_day)
    # 获取星期几（0为星期一，1为星期二，以此类推）
    s_beginWeekDay = dateObject.weekday() + 1
    print("开始考试星期几--", s_beginWeekDay)
    
    resultFile = os.path.splitext(s_originalFile)[0] + "_result.xlsx"
    sortPersons(personArr)
    generateResult(headRow, personArr, resultFile)
    print(u"导出完毕：{}".format(resultFile))

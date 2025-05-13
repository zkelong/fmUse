# coding=utf-8
import sys, os
# 解析 xls
import xlrd
# 解析 xlsx
import openpyxl
import math
from datetime import datetime, timedelta

# ///////////////////////////////////////////////////
# 排布规则:
# 能排一天的不跨天
# 一周能考完的不跨周
# openpyxl: pip install openpyxl
# 使用 python3 运行脚本
# 每个人还是连考，每次考一科
# 标记为 1 的优先放入 78 座教室，如果放不下，放入 86 座教室
# 标记为 2 的优先放入 86 座教室，放不下再放 1 的
# 6月14.15.16号
# ///////////////////////////////////////////////////

# 每日场次 python 2.7 整数相除为整数，没有小数
# 使用 python3
s_dayNum = 5
# 每场人数
s_sessionNum = 20
# 周末考试
s_weekEnd = False

# 时间排布
s_begin_year = 2023
s_begin_month = 6
s_begin_day = 14
s_times = ["8:30", "10:30", "12:30", "14:30", "16:30", "18:30"]

# 表结构
# 考点编号*,考点名称*,时间单元编号*,考试开始时间*,考场编号*,考场名称*,座位号*,考生学号*,考生姓名*,课程编号*,课程名称*,试卷号*
s_keyIndex = 0  # 索引字段-学号
s_typeIndex = 6  # 类型索引-课程编号

# 教室标记
s_roomTypeIndex = 7

# 开始时间(周几)
s_beginWeekDay = 0
# 开始周考试天数
s_beginWeekSessionNum = 0
# 表头
s_head = ""


# 待处理文件
s_originalFile = None

# 考生
s_personKeys = []


# 获取指定年月日相差天数的，年月日
def calculate_new_date(year, month, day, days_difference):
    # 构造日期对象
    current_date = datetime(year, month, day)
    # 计算相差指定天数后的日期
    new_date = current_date + timedelta(days=days_difference)
    return new_date.year, new_date.month, new_date.day

# 人
def getTime(session):
    global s_begin_year
    global s_begin_month
    global s_begin_day
    global s_dayNum
    global s_times

    year, month, day = calculate_new_date(s_begin_year, s_begin_month, s_begin_day, (session - 1) / s_dayNum)
    
    index = session % s_dayNum
    if index == 0:
        index = len(s_times)
    index = index - 1

    if month < 10:
        month = "0" + str(month)
    if day < 10:
        day = "0" + str(day)
    return str(year) + "/" + str(month) + "/" + str(day) + " " + s_times[int(index)], str(session)

class Person:
    def __init__(self, _key, _type, _roomType):
        self.m_key = _key
        self.m_type = _type
        self.m_roomType = _roomType
        self.rows = []

    def getKey(self):
        return self.m_key
    
    def getRoomType(self):
        return self.m_roomType

    def addRow(self, row, _type):
        self.rows.append(row)

    def getRowNum(self):
        return len(self.rows)

    def getRows(self):
        return self.rows

    def setSession(self, session):
        for row in self.rows:
            time, tag = getTime(session)
            row.addValues(session, time, tag)
            session += 1


# 寻找已有的 person
def getPerson(_key, personArr):
    for person in personArr:
        if person.getKey() == _key:
            return person
    return None


# 获取行的 key, type
def getKeyAndType(row):
    global s_keyIndex
    global s_typeIndex
    global s_roomTypeIndex

    return row.values[s_keyIndex], row.values[s_typeIndex], row.values[s_roomTypeIndex]


# ///////// 读 excel 内容 ///////////////
# excel 行数据
class ExcelRow:
    # row xlsx 读得的值
    def __init__(self, row = None):
        self.values = []
        if row:
            for cell in row:
                self.values.append(cell.value)
        # print(row[0].value)

    # xls 读取的值
    def addValue(self, value):
        # 数字单元格的结果后面多了一个.0，这是因为xlrd默认将数字单元格的数据类型解析为浮点数（float），即使实际上是整数
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        self.values.append(value)

    # 加：场次，时间，时间编号
    def addValues(self, session, time, tag):
        self.values.append(session)
        self.values.append(time)
        self.values.append(tag)

    def getRowContent(self):
        content = ""
        for value in self.values:
            content += str(value) + ","
        return content.replace(",", "")


# 读原文件 xls
def readXls():
    global s_originalFile

    headRow = None
    personArr = []
    _workbook = xlrd.open_workbook(s_originalFile)
    _sheets = _workbook.sheets()
    if len(_sheets) == 0:
        print(u"excel 文件内容为空！！")
        exit()
    if len(_sheets) > 1:
        print(u"警告！！存在多个sheet，只处理第一个sheet，如后面的 sheet 要处理，都拷贝到第一个 sheet!!!")
        print(u"警告！！存在多个sheet，只处理第一个sheet，如后面的 sheet 要处理，都拷贝到第一个 sheet!!!")
        print(u"警告！！存在多个sheet，只处理第一个sheet，如后面的 sheet 要处理，都拷贝到第一个 sheet!!!")
    # sheet = _workbook.sheet_by_index(0)
    sheet = _sheets[0]
    row_index = 0
    for row in range(sheet.nrows):
        if row_index == 0:
            headRow = ExcelRow()
            for col in range(sheet.ncols):
                value = sheet.cell_value(row, col)
                headRow.addValue(value)
            headRow.addValues("session", "time", "time_tag")
        else:
            excel_row = ExcelRow()
            for col in range(sheet.ncols):
                value = sheet.cell_value(row, col)
                excel_row.addValue(value)
            _key, _type, _rommType = getKeyAndType(excel_row)
            person = getPerson(_key, personArr)
            if person is None:
                s_personKeys.append(_key)
                person = Person(_key, _type, _rommType)
                personArr.append(person)
            person.addRow(excel_row, _type, _rommType)
        row_index += 1

    print(u"场数：{}".format(row_index - 1))
    print(u"人数：{}".format(len(personArr)))
    return headRow, personArr


# 读原文件 xlsx
def readXlsx():
    global s_originalFile

    headRow = None
    personArr = []
    workbook = openpyxl.load_workbook(s_originalFile)
    if len(workbook.sheetnames) == 0:
        print(u"excel 文件内容为空！！")
        exit()
    elif len(workbook.sheetnames) > 1:
        print(u"警告！！存在多个sheet，只处理第一个sheet，如后面的 sheet 要处理，都拷贝到第一个 sheet!!!")
        print(u"警告！！存在多个sheet，只处理第一个sheet，如后面的 sheet 要处理，都拷贝到第一个 sheet!!!")
        print(u"警告！！存在多个sheet，只处理第一个sheet，如后面的 sheet 要处理，都拷贝到第一个 sheet!!!")
    sheet = workbook.worksheets[0]
    row_index = 0
    for _row in sheet.iter_rows():
        if row_index == 0:
            headRow = ExcelRow(_row)
            headRow.addValues("session", "time", "time_tag")
        else:
            row = ExcelRow(_row)
            _key, _type, _rommType = getKeyAndType(row)
            person = getPerson(_key, personArr)
            if person is None:
                s_personKeys.append(_key)
                person = Person(_key, _type, _rommType)
                personArr.append(person)
            person.addRow(row, _type)
        row_index += 1

    print(u"场数：{}".format(row_index - 1))
    print(u"人数：{}".format(len(personArr)))
    return headRow, personArr


# 场次多的人排前面
def sortPersons(personArr):
    print(u"场次多的排前面")
    length = len(personArr)
    for i in range(0, length):
        for j in range(i + 1, length):
            if personArr[j].getRowNum() > personArr[i].getRowNum():
                temp = personArr[i]
                personArr[i] = personArr[j]
                personArr[j] = temp
    print(u"排序完毕")


# 周内开始，并周末不安排考试时，返回第一周能进行的场次
def setFirstWeekExamNum():
    global s_dayNum
    global s_beginWeekDay
    global s_weekEnd
    global s_beginWeekSessionNum

    # 周末考试 或 从周一开始考，则不考虑一周无法考完的情况
    if s_weekEnd or s_beginWeekDay == 1:
        s_beginWeekSessionNum = 0
    s_beginWeekSessionNum = (5 - s_beginWeekDay + 1) * s_dayNum


# 安排场次是否超天或超周
def isSessionOverDayOrWeek(session, person):
    global s_dayNum
    
    _personLineNum = person.getRowNum()
    ### 是否超天
    # 需要用到考试天数
    _needDay = math.ceil(_personLineNum / s_dayNum)
    # 耗费天数
    _firstDayLeft = 0  # 第一天剩余场次
    if session < s_dayNum:
        _firstDayLeft = s_dayNum - session + 1
    else:
        if session % s_dayNum == 0:
            _firstDayLeft = 1
        else:
            _firstDayLeft = s_dayNum - (session % s_dayNum) + 1

    if _personLineNum <= _firstDayLeft:
        return False
    _costDay = math.ceil((_personLineNum - _firstDayLeft) / s_dayNum) + 1

    if _costDay > _needDay:
        return True
    return False

# 获取当前场次-教室1
def getNowSession1(sessionInfo1, sessionInfo2, person):
    _session = 0
    _roomTypes = []
    while True:
        _session += 1
        # 只有3天，共18场
        if _session > 18:
            break
        # 教室1 座位：78
        if str(_session) in sessionInfo1 and sessionInfo1[str(_session)] >= 78:
            # 当前场次已满
            continue
        if isSessionOverDayOrWeek(_session, person):
            # 超天或超周
            continue
        if not isNextSessionEnough(_session, sessionInfo1, person, 78):
            # 后续 session 排满了
            continue
        return _session, _roomTypes
    _session = 0
    while True:
        _session += 1
        # 只有3天，共18场
        if _session > 18:
            break
        if str(_session) in sessionInfo1 and sessionInfo1[str(_session)] >= 78:
            if str(_session) in sessionInfo2 and sessionInfo2[str(_session)] >= 86:
                # 当前场次已满
                continue
        if isSessionOverDayOrWeek(_session, person):
            # 超天或超周
            continue
        if not isNextSessionEnough(_session, sessionInfo1, person, 78):
            if not isNextSessionEnough(_session, sessionInfo2, person, 86):
                # 后续 session 排满了
                continue
        rowNum = person.getRowNum()
        for i in range(_session, _session + person.getRowNum()):
            _sKey = str(i)
            if _sKey in sessionInfo1 and sessionInfo1[_sKey] >= 78:
                _roomTypes.append(2)
            else:
                _roomTypes.append(1)
        return _session
    return None, None

# 教室标记
def separatePerson(personArr):
    result = []
    resultSpecial = []
    for p in personArr:
        if p.getRoomType() == 1:
            result.append(p)
        else:
            resultSpecial.append(p)
    return result, resultSpecial

# 后续场次是否足够
def isNextSessionEnough(session, sessionInfo, person, num):
    for i in range(session, session + person.getRowNum()):
        _sKey = str(i)
        if _sKey in sessionInfo and sessionInfo[_sKey] >= num:
            return False
    return True


# 记录 session 数量
def addSessionInfo(sessionInfo, session, person):
    for i in range(session, session + person.getRowNum()):
        _sKey = str(i)
        if _sKey not in sessionInfo:
            sessionInfo[_sKey] = 1
        else:
            sessionInfo[_sKey] += 1


# 开始排序，获得结果
def generateResult(headRow, personArr, specailPersonArr, resultFile):
    global s_head

    print(u"获取结果内容")
    # 教室1场次信息：{"1":100} 场次对应次数
    _sessionInfo1 = {}
    # 教室2场次信息：{"1":100} 场次对应次数
    _sessionInfo2 = {}
    setFirstWeekExamNum()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    print(u"写入头.....")
    for index, value in enumerate(headRow.values):
        # row, column 从 1 开始
        cell = sheet.cell(row=1, column=index + 1)
        cell.value = value

    print(u"写入详细.....")
    # 上面加了头，从第二行开始写
    _rowIndex = 2
    personIndex = 0
    
    
    for person in personArr:
        # 获取从哪场开始排
        session = getNowSession1(_sessionInfo1, _sessionInfo2, person)
        if session is None:
            print("安排不下")
            exit()
        addSessionInfo(_sessionInfo, session, person)
        person.setSession(session)
        personIndex += 1
        for index, row in enumerate(person.getRows()):
            for _index, value in enumerate(row.values):
                # row, column 从 1 开始
                cell = sheet.cell(row=_rowIndex, column=_index + 1)
                # 科学计数法问题
                # 设置单元格格式
                cell.number_format = '0'  # 或者使用 '0.00' 等形式，确保数字以常规格式显示，而非科学计数法
                cell.value = value
            _rowIndex += 1
            
    for person in specailPersonArr:
        # 获取从哪场开始排
        session = getNowSession(_sessionInfo, person)
        addSessionInfo(_sessionInfo, session, person)
        person.setSession(session)
        personIndex += 1
        for index, row in enumerate(person.getRows()):
            for _index, value in enumerate(row.values):
                # row, column 从 1 开始
                cell = sheet.cell(row=_rowIndex, column=_index + 1)
                # 科学计数法问题
                # 设置单元格格式
                cell.number_format = '0'  # 或者使用 '0.00' 等形式，确保数字以常规格式显示，而非科学计数法
                cell.value = value
            _rowIndex += 1

    workbook.save(resultFile)
    workbook.close()


if __name__ == '__main__':
    print("__main__>>>>>>>>>>>>>>>>>>>>>>>>")
    if sys.version_info.major < 3:
        print(u"使用 python3 执行脚本！！！！！")
        exit()
    
    if len(sys.argv) < 2:
        print(u"请输入要排序的文件名!")
        exit()
    s_originalFile = sys.argv[1]
    if not os.path.isfile(s_originalFile):
        print(u"输入的文件不存在！")
        exit()
        
    headRow = None
    personArr = None
    _, file_ext = os.path.splitext(s_originalFile)
    isXlsx = False
    if file_ext == ".xlsx":
        headRow, personArr = readXlsx()
    elif file_ext == ".xls":
        headRow, personArr = readXls()
    else:
        print(u"只支持排序 .xlsx 和 .xls 文件")
        exit()

    # 开始时间(周几)
    dateObject = datetime(s_begin_year, s_begin_month, s_begin_day)
    # 获取星期几（0为星期一，1为星期二，以此类推）
    s_beginWeekDay = dateObject.weekday() + 1
    # 分离特殊角色
    personArr, specailPersonArr= separatePerson(personArr)
    resultFile = os.path.splitext(s_originalFile)[0] + "_result.xlsx"
    sortPersons(personArr)
    sortPersons(specailPersonArr)
    generateResult(headRow, personArr, specailPersonArr, resultFile)
    print(u"导出完毕：{}".format(resultFile))

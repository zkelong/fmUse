# coding=utf-8
from tools import GetExcelDatas
# 解析 xls
import xlrd
# 解析 xlsx
import openpyxl

# ///////////////////////////////////////////////////
# 获取每个时间段的考试科目人数
# ///////////////////////////////////////////////////

# 试卷号排序
g_sort = [
    22542,22617,23617,23780,22194,22208,22508,11073,11568,11171,23722,22109,22332,11646,24186,11141,22402,22412,22099,22528,22019,11067,22247,22098,23980,11505,24107,24107,22202,24108,11054,22175,11334,22097,22251,22666,11511,22175,22691,22180,24154,24176,23990,11258,22517,22006,11477,11575,22110,22238,22094,11080,24149,22108,22409,24046,22072,22096,22246,22107,11620,22223,22320,22668,24153
]

############ 名单表 ############
# 表头行数
g_head_lines = 1
# 统计数据 日期	考场号	试卷号	试卷名称	考场人数	教室
# 日期列标
g_date_index = 11
# 考场号列标
g_exam_index = 6
# 试卷号列标
g_exam_paper_index = 13
# 试卷名称列标
g_exam_name_index = 14
# 教室列表
g_room_index = 19
# 学号
g_number_index = 0

# 错误信息
g_error_msg = ""

# 检查试卷号排序是否有重合的,分两场的
def checkSortReapt():
    global g_sort
    
    repeat = []
    repeatStr = ""
    for i, value in enumerate(g_sort):
        for _i, _value in enumerate(g_sort):
            if i != _i and value == _value:
                if value not in repeat:
                    repeat.append(value)
                    repeatStr += str(_value) + ","
    return repeatStr

# 考试场次信息
class ExamInfo:
    def __init__(self, row):
        # 日期
        self.date = row.getValue(g_date_index)
        # 考场号
        self.exam = row.getValue(g_exam_index).strip()
        # 试卷号
        self.examPaper = row.getValue(g_exam_paper_index).strip()
        # 试卷名称
        self.examName = row.getValue(g_exam_name_index)
        # 教室
        self.room = row.getValue(g_room_index)
        # 包含的人
        self.roles = [row.getValue(g_number_index)]
        # 人数
        self.num = 1
    
    def addNum(self):
        self.num += 1

# 找到已有的场次信息
def findExamInfo(exams, row):
    global g_exam_index
    
    for e in exams:
        if e.exam == row.getValue(g_exam_index).strip():
            return e
    return None

# 是否已有场次信息
def addToExams(exams, row):
    global g_exam_index
    global g_error_msg
    exam = findExamInfo(exams, row)
    if exam is None:
        exams.append(ExamInfo(row))
        return
    role = row.getValue(g_number_index)
    if role in exam.roles:
        g_error_msg += str(exam.exam) + "已有" + str(role) + ";"
    else:
        exam.roles.append(role)
    exam.addNum()
    return True

def writeExelRow(sheet, rowIndex, info):
    # ["日期","考场号","试卷号","试卷名称","考场人数","教室"]
    values = [info.date, info.exam, info.examPaper, info.examName, info.room, info.num]
    for _index, _value in enumerate(values):
        # row, column 从 1 开始
        cell = sheet.cell(row=rowIndex, column=_index + 1)
        # 科学计数法问题
        # 设置单元格格式
        cell.number_format = '0'  # 或者使用 '0.00' 等形式，确保数字以常规格式显示，而非科学计数法
        cell.value = _value

# 所有考试信息
class AllExamInfo:
    def __init__(self):
        self.file = ""
        
    # 根据 考试名单 获取每场考试信息
    def getExamInfos():
        # 关注的信息
        head = ["试卷号","试卷名称","考场人数","考场号","教室","日期"]
        # 关注信息对应的列
        headIndexs = [13, 14, 0, 6, 19, 11]
        
        # 统计数据 日期	考场号	试卷号	试卷名称	考场人数	教室
# 日期列标
g_date_index = 11
# 考场号列标
g_exam_index = 6
# 试卷号列标
g_exam_paper_index = 13
# 试卷名称列标
g_exam_name_index = 14
# 教室列表
g_room_index = 19
# 学号
g_number_index = 0


if __name__ == '__main__':
    print("__main__>>>>>>>>>>>>>>>>>>>>>>>>")
    repeat = checkSortReapt()
    if len(repeat) > 0:
        print("分为两场的考试:" + repeat)
        
    rowArr = GetExcelDatas.getExcelData("简阳考试名单.xlsx", True)
    
    exams = []
    index = 0
    for row in rowArr:
        addToExams(exams, row)

    if len(g_error_msg) > 0:
        print(g_error_msg)
        exit()

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    print(u"写入头.....")
    for _index, _value in enumerate(["日期","考场号","试卷号","试卷名称","考场人数","教室"]):
        # row, column 从 1 开始
        cell = sheet.cell(row=1, column=_index + 1)
        cell.value = _value
        
    print(u"写入详细.....")
    # 上面加了头，从第二行开始写
    _rowIndex = 2
    findExamPaper = []
    unFindExamPaper = []
    unFindStr = ""
    # 根据指定顺序排序
    for s in g_sort:
        find = False
        for e in exams:
            if str(s) == str(e.examPaper):
                findExamPaper.append(str(s))
                find = True
                writeExelRow(sheet, _rowIndex, e)
                _rowIndex += 1
        if not find:
            unFindExamPaper.append(str(s))
            unFindStr += str(s) + ","
    
    if len(unFindStr) > 0:
        print("排序中未找到的试卷号：" + unFindStr)
        
    notInSortStr = ""      
    for e in exams:
        if str(e.examPaper) not in findExamPaper:
            writeExelRow(sheet, _rowIndex, e)
            _rowIndex += 1
            notInSortStr += str(e.examPaper) + ","
            
    if len(notInSortStr) > 0:
        print("不在排序中的试卷号：" + notInSortStr)
    
    workbook.save("result.xlsx")
    workbook.close()
    print("end=========================")
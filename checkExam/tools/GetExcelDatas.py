# coding=utf-8
import sys, os
# 解析 xls
import xlrd
# 解析 xlsx
import openpyxl
import math

# ///////////////////////////////////////////////////
# 获取 excel 内容
# ///////////////////////////////////////////////////
# excel 行数据
class ExcelRow:
    # row xlsx 读得的值
    def __init__(self, row = None, isHead = False):
        self.isHead = isHead
        self.values = []
        if row:
            for cell in row:
                self.values.append(cell.value)

    # xls 读取的值
    def addValue(self, value):
        # 数字单元格的结果后面多了一个.0, 这是因为xlrd默认将数字单元格的数据类型解析为浮点数（float）, 即使实际上是整数
        if isinstance(value, float) and value.is_integer():
            value = int(value)
        self.values.append(value)
    
    # 获取列数据
    def getValue(self, index):
        return self.values[index]

# 读原文件 xls
def readXls(file, haveHead):
    _workbook = xlrd.open_workbook(file)
    _sheets = _workbook.sheets()
    if len(_sheets) == 0:
        print(u"excel 文件内容为空!!")
        exit()
    if len(_sheets) > 1:
        print(u"警告!!存在多个sheet, 只处理第一个sheet, 如后面的 sheet 要处理, 都拷贝到第一个 sheet!!!")
        print(u"警告!!存在多个sheet, 只处理第一个sheet, 如后面的 sheet 要处理, 都拷贝到第一个 sheet!!!")
        print(u"警告!!存在多个sheet, 只处理第一个sheet, 如后面的 sheet 要处理, 都拷贝到第一个 sheet!!!")
    # sheet = _workbook.sheet_by_index(0)
    sheet = _sheets[0]
    row_index = 0
    rowArr = []
    for row in range(sheet.nrows):
        rowArr.append(ExcelRow(isHead = haveHead and row_index == 0))
        row_index += 1
    return row


# 读原文件 xlsx
def readXlsx(file, haveHead):
    workbook = openpyxl.load_workbook(file)
    if len(workbook.sheetnames) == 0:
        print(u"文件:{} 内容为空!!".format(file))
        exit()
    elif len(workbook.sheetnames) > 1:
        print(u"警告!!存在多个sheet, 只处理第一个sheet, 如后面的 sheet 要处理, 都拷贝到第一个 sheet!!!")
        print(u"警告!!存在多个sheet, 只处理第一个sheet, 如后面的 sheet 要处理, 都拷贝到第一个 sheet!!!")
        print(u"警告!!存在多个sheet, 只处理第一个sheet, 如后面的 sheet 要处理, 都拷贝到第一个 sheet!!!")
    sheet = workbook.worksheets[0]
    rowArr = []
    row_index = 0
    for _row in sheet.iter_rows():
        rowArr.append(ExcelRow(_row, haveHead and row_index == 0))
        row_index += 1
    return rowArr

def getExcelData(file, haveHead=True):
    if not os.path.isfile(file):
        print(u"文件:{} 不存在!".format(file))
        exit()
    rowArr = None
    _, file_ext = os.path.splitext(file)
    if file_ext == ".xlsx":
        rowArr = readXlsx(file, haveHead)
    elif file_ext == ".xls":
        rowArr = readXls(file, haveHead)
    else:
        print(u"只支持获取 .xlsx 和 .xls 文件内容")
        exit()
    return rowArr

# test
rowArr = getExcelData("简阳考试名单.xlsx")
print(len(rowArr))